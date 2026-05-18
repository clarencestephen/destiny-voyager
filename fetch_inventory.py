"""
fetch_inventory.py
==================
Pull your full Destiny 2 inventory (vault + per-character equipped + character
inventories + in-game saved loadouts) via Bungie OAuth, then write two new
sheets to your workbook:

  INVENTORY     — every item you own, sortable by class/element/tier/tag
  MY LOADOUTS   — in-game saved loadouts (Lightfall+ native feature) +
                  any custom loadouts you've saved with this toolkit

Usage:
    python3 fetch_inventory.py            # signs in if needed, then fetches
    python3 fetch_inventory.py --tag-only # skip API fetch, just refresh tag column

Tags live in user_config.json under `item_tags` keyed by instance ID.
Tag values: favorite | keep | infuse | junk | archive | (blank)
"""

import argparse
import json
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CONFIG_PATH = Path("user_config.json")

# Reference: manifest is cached by decode_dim.py — we reuse that cache
MANIFEST_CACHE = Path("./manifest_cache")
MANIFEST_TABLE = "DestinyInventoryItemDefinition"

CLASS_NAMES = {0: "Titan", 1: "Hunter", 2: "Warlock", 3: "Unknown"}

# DamageType → readable
DAMAGE_TYPES = {
    0: "None", 1: "Kinetic", 2: "Arc", 3: "Solar", 4: "Void",
    6: "Stasis", 7: "Strand",
}

# Item tier (rarity)
TIER_NAMES = {
    2: "Basic", 3: "Common", 4: "Rare", 5: "Legendary", 6: "Exotic",
}

# Standard slot buckets we care about
SLOT_BUCKETS = {
    1498876634: "Kinetic",
    2465295065: "Energy",
    953998645:  "Heavy",
    3448274439: "Helmet",
    3551918588: "Gauntlets",
    14239492:   "Chest Armor",
    20886954:   "Leg Armor",
    1585787867: "Class Armor",
    4023194814: "Ghost",
    284967655:  "Ship",
    2025709351: "Sparrow",
    3284755031: "Subclass",
}

VALID_TAGS = ["favorite", "keep", "infuse", "junk", "archive"]


def load_config():
    if not CONFIG_PATH.exists():
        sys.exit("ERROR: user_config.json not found. Run setup.py first.")
    return json.loads(CONFIG_PATH.read_text())


def save_config(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2) + "\n")
    try:
        CONFIG_PATH.chmod(0o600)
    except Exception:
        pass


def load_manifest():
    """Load the cached item definitions. Run decode_dim.py first if missing."""
    f = MANIFEST_CACHE / f"{MANIFEST_TABLE}.json"
    if not f.exists():
        sys.exit(
            "ERROR: manifest not cached. Run `python3 decode_dim.py` once first "
            "to populate manifest_cache/. (It only needs to fetch the manifest once.)"
        )
    print(f"  Loading manifest from cache...")
    return json.loads(f.read_text())


def item_defs_for(items_raw, manifest):
    """Resolve raw inventory items → enriched dicts with names, types, etc."""
    out = []
    for it in items_raw:
        h = str(it.get("itemHash"))
        defn = manifest.get(h, {})
        out.append({
            "instanceId": it.get("itemInstanceId"),
            "hash": h,
            "name": defn.get("displayProperties", {}).get("name", f"?({h})"),
            "tier": TIER_NAMES.get(defn.get("inventory", {}).get("tierType", 0), ""),
            "type": defn.get("itemTypeDisplayName", ""),
            "class": CLASS_NAMES.get(defn.get("classType", 3), "Any"),
            "element": DAMAGE_TYPES.get(defn.get("defaultDamageType", 0), ""),
            "slot": SLOT_BUCKETS.get(defn.get("inventory", {}).get("bucketTypeHash", 0),
                                     "Other"),
            "bucket_hash": defn.get("inventory", {}).get("bucketTypeHash", 0),
            "raw_quantity": it.get("quantity", 1),
        })
    return out


def collect_inventory(snapshot, manifest):
    """
    Walk a profile snapshot and return a flat list of items with their location
    (vault / character N) + instance data (power level when available).
    """
    items = []
    prof = snapshot["profile"]

    # Vault items
    vault_items = (prof.get("profileInventory", {}).get("data", {})
                   .get("items", []))
    for it in item_defs_for(vault_items, manifest):
        it["location"] = "Vault"
        items.append(it)

    # Per-character inventory + equipped
    char_inv = prof.get("characterInventories", {}).get("data", {})
    equipped = prof.get("characterEquipment", {}).get("data", {})
    chars = prof.get("characters", {}).get("data", {})

    for char_id, c in chars.items():
        cls = CLASS_NAMES.get(c.get("classType", 3), "Any")
        light = c.get("light", 0)
        # Inventory for this character
        for it in item_defs_for(char_inv.get(char_id, {}).get("items", []), manifest):
            it["location"] = f"{cls} (Light {light})"
            items.append(it)
        # Equipped on this character
        for it in item_defs_for(equipped.get(char_id, {}).get("items", []), manifest):
            it["location"] = f"{cls} EQUIPPED"
            items.append(it)

    # Enrich with itemInstances component (power level)
    instances = prof.get("itemComponents", {}).get("instances", {}).get("data", {})
    for it in items:
        inst = instances.get(str(it["instanceId"]), {}) if it["instanceId"] else {}
        it["power"] = inst.get("primaryStat", {}).get("value")
    return items


def collect_loadouts(snapshot, manifest):
    """In-game saved loadouts per character (Lightfall+)."""
    prof = snapshot["profile"]
    chars = prof.get("characters", {}).get("data", {})
    loadout_data = prof.get("characterLoadouts", {}).get("data", {})
    out = []
    for char_id, c in chars.items():
        cls = CLASS_NAMES.get(c.get("classType", 3), "Any")
        loadouts = loadout_data.get(char_id, {}).get("loadouts", [])
        for idx, ld in enumerate(loadouts, 1):
            name_hash = str(ld.get("nameHash", 0))
            color_hash = str(ld.get("colorHash", 0))
            icon_hash = str(ld.get("iconHash", 0))
            # If all hashes are 0 the slot is empty
            if name_hash == "0" and not ld.get("items"):
                continue
            items = []
            for slot_item in ld.get("items", []):
                ih = str(slot_item.get("itemInstanceId"))
                items.append(ih)
            out.append({
                "class": cls,
                "slot": idx,
                "name_hash": name_hash,
                "color_hash": color_hash,
                "icon_hash": icon_hash,
                "item_count": len(ld.get("items", [])),
                "raw": ld,
            })
    return out


# ---------- Workbook writers ----------

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def setc(ws, ref, val, *, bold=False, size=11, color="000000",
         fill=None, italic=False, align="left", wrap=True):
    c = ws[ref]
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER


def write_inventory_sheet(workbook_path, items, tags):
    """Rewrite the INVENTORY sheet from scratch."""
    wb = load_workbook(workbook_path)
    if "INVENTORY" in wb.sheetnames:
        del wb["INVENTORY"]
    ws = wb.create_sheet("INVENTORY", 4)  # insert after WISHLIST
    widths = [4, 28, 12, 12, 22, 20, 8, 10, 18, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:J1")
    setc(ws, "A1",
         f"  INVENTORY — {len(items):,} items  (fetched {time.strftime('%Y-%m-%d %H:%M')})",
         bold=True, size=14, color="FFFFFF", fill="1F2937")
    ws.row_dimensions[1].height = 26

    headers = ["#", "Name", "Tier", "Element", "Type", "Slot", "Power",
               "Tag", "Location", "Instance ID"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}3", h, bold=True, fill="F3F4F6",
             align="center" if h in ("#", "Power", "Tag") else "left")

    row = 4
    for i, it in enumerate(items, 1):
        tag = tags.get(str(it["instanceId"]), "") if it["instanceId"] else ""
        tag_color = {
            "favorite": "FCD34D",
            "keep":     "A7F3D0",
            "infuse":   "FDE68A",
            "junk":     "FCA5A5",
            "archive":  "E5E7EB",
        }.get(tag, None)

        setc(ws, f"A{row}", i, align="center", size=9, color="6B7280")
        setc(ws, f"B{row}", it["name"], bold=(it["tier"] == "Exotic"))
        setc(ws, f"C{row}", it["tier"], size=10,
             color="F59E0B" if it["tier"] == "Exotic" else "000000")
        setc(ws, f"D{row}", it["element"], size=10)
        setc(ws, f"E{row}", it["type"], size=10)
        setc(ws, f"F{row}", it["slot"], size=10)
        setc(ws, f"G{row}", it.get("power") or "", align="center", size=10)
        setc(ws, f"H{row}", tag, align="center", size=10, bold=bool(tag),
             fill=tag_color)
        setc(ws, f"I{row}", it["location"], size=9, color="6B7280")
        setc(ws, f"J{row}", str(it["instanceId"] or ""), size=8, color="9CA3AF")
        row += 1

    ws.freeze_panes = "A4"  # keep header visible when scrolling
    wb.save(workbook_path)
    print(f"  ✓ Wrote INVENTORY sheet ({len(items):,} rows)")


def write_loadouts_sheet(workbook_path, loadouts):
    wb = load_workbook(workbook_path)
    if "MY LOADOUTS" in wb.sheetnames:
        del wb["MY LOADOUTS"]
    ws = wb.create_sheet("MY LOADOUTS", 5)
    widths = [4, 14, 8, 14, 10, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    setc(ws, "A1", f"  MY LOADOUTS — in-game saved + custom (Lightfall+ native)",
         bold=True, size=14, color="FFFFFF", fill="1F2937")
    ws.row_dimensions[1].height = 26

    headers = ["#", "Class", "Slot", "Name Hash", "Items", "Notes"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}3", h, bold=True, fill="F3F4F6",
             align="center" if h in ("#", "Slot", "Items") else "left")

    row = 4
    for i, ld in enumerate(loadouts, 1):
        setc(ws, f"A{row}", i, align="center")
        setc(ws, f"B{row}", ld["class"], bold=True,
             color={"Hunter": "2563EB", "Titan": "DC2626",
                    "Warlock": "7C3AED"}.get(ld["class"], "000000"))
        setc(ws, f"C{row}", ld["slot"], align="center")
        setc(ws, f"D{row}", ld["name_hash"], size=9, color="6B7280")
        setc(ws, f"E{row}", ld["item_count"], align="center")
        setc(ws, f"F{row}", "(item names resolve from instance IDs in INVENTORY)",
             italic=True, size=9, color="6B7280")
        row += 1

    if not loadouts:
        ws.merge_cells(f"A4:F4")
        setc(ws, "A4",
             "  No in-game loadouts found. Save some in-game (post-Lightfall feature) "
             "and re-run fetch_inventory.py.",
             italic=True, size=10, color="6B7280", fill="F9FAFB")

    ws.freeze_panes = "A4"
    wb.save(workbook_path)
    print(f"  ✓ Wrote MY LOADOUTS sheet ({len(loadouts)} rows)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tag-only", action="store_true",
                    help="Skip API fetch, just refresh the Tag column from user_config")
    args = ap.parse_args()

    cfg = load_config()
    workbook_path = cfg.get("workbook_path", "./my_loadouts.xlsx")
    if not Path(workbook_path).exists():
        sys.exit(f"ERROR: workbook not found at {workbook_path}")

    tags = cfg.get("item_tags", {})

    if args.tag_only:
        # Re-write just the Tag column — would require re-reading existing rows.
        # For simplicity, full re-fetch is recommended.
        print("  --tag-only not yet supported. Run without flag for full refresh.")
        return

    # Ensure signed in
    from auth import ensure_signed_in
    ensure_signed_in(client_id=cfg.get("oauth_client_id", "52250"))

    from bungie_client import BungieClient
    client = BungieClient()
    print("  Fetching inventory snapshot from Bungie API...")
    snapshot = client.get_inventory_snapshot()
    print(f"  Membership: {snapshot['membership']['displayName']} "
          f"(type {snapshot['membership']['membershipType']})")

    manifest = load_manifest()
    print("  Resolving items + locations...")
    items = collect_inventory(snapshot, manifest)
    loadouts = collect_loadouts(snapshot, manifest)
    print(f"  Found {len(items):,} items, {len(loadouts)} saved loadouts")

    print("  Writing workbook sheets...")
    write_inventory_sheet(workbook_path, items, tags)
    write_loadouts_sheet(workbook_path, loadouts)
    print(f"  Done. Open {workbook_path} → INVENTORY / MY LOADOUTS sheets.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n  Aborted.")
        sys.exit(130)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
