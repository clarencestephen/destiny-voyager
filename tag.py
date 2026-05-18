"""
tag.py
======
Tag items in your inventory. Tags persist in user_config.json under `item_tags`,
keyed by Destiny item instance ID.

Tag values (DIM-compatible):
  favorite   gold star — your go-to gear, never delete
  keep       green — solid roll, keep around
  infuse     yellow — for power infusion fuel
  junk       red — flagged to dismantle
  archive    grey — out of meta but you want to remember

Usage:
    # Tag by name (partial-match search) — interactive picker
    python3 tag.py "fatebringer"

    # Tag directly by instance ID
    python3 tag.py --id 6917530000000001234 --tag favorite

    # List all tagged items
    python3 tag.py --list

    # Remove tag
    python3 tag.py --id <id> --untag

    # Clear all tags
    python3 tag.py --clear-all

After tagging, re-run `python3 fetch_inventory.py` to update the INVENTORY sheet
(or wait until the next refresh).
"""

import argparse
import json
import sys
from pathlib import Path

CONFIG_PATH = Path("user_config.json")
VALID_TAGS = ["favorite", "keep", "infuse", "junk", "archive"]


def load_cfg():
    if not CONFIG_PATH.exists():
        sys.exit("ERROR: user_config.json not found.")
    return json.loads(CONFIG_PATH.read_text())


def save_cfg(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2) + "\n")
    try:
        CONFIG_PATH.chmod(0o600)
    except Exception:
        pass


def search_inventory(query):
    """
    Search the most-recent INVENTORY sheet for items matching `query`.
    Returns list of (instance_id, name, tier, location).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: openpyxl not installed.")
    cfg = load_cfg()
    wb_path = cfg.get("workbook_path", "./my_loadouts.xlsx")
    if not Path(wb_path).exists():
        sys.exit(f"ERROR: {wb_path} not found. Run fetch_inventory.py first.")
    wb = load_workbook(wb_path, read_only=True)
    if "INVENTORY" not in wb.sheetnames:
        sys.exit("ERROR: no INVENTORY sheet yet. Run fetch_inventory.py first.")
    ws = wb["INVENTORY"]
    matches = []
    q = query.lower()
    # Cols: A=#, B=Name, C=Tier, D=Element, E=Type, F=Slot, G=Power, H=Tag, I=Location, J=InstanceId
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[1]:
            continue
        name = str(row[1])
        if q in name.lower():
            matches.append({
                "instance_id": str(row[9]) if row[9] else "",
                "name": name,
                "tier": row[2] or "",
                "element": row[3] or "",
                "power": row[6] or "",
                "location": row[8] or "",
            })
    return matches


def cmd_search(args, cfg):
    matches = search_inventory(args.query)
    if not matches:
        print(f"  No items match {args.query!r}")
        return
    print(f"  {len(matches)} match(es) for {args.query!r}:")
    for i, m in enumerate(matches[:20], 1):
        existing = cfg.get("item_tags", {}).get(m["instance_id"], "")
        tag_str = f" [{existing}]" if existing else ""
        print(f"    {i:3d}.  {m['name']:32}  {m['tier']:10}  pow {m['power']:>4}  "
              f"{m['location']:24}  {m['instance_id']}{tag_str}")
    if len(matches) > 20:
        print(f"    ... +{len(matches) - 20} more (narrow your query)")

    if args.tag:
        if args.tag not in VALID_TAGS:
            sys.exit(f"  Invalid tag. Use one of: {', '.join(VALID_TAGS)}")
        if len(matches) == 1:
            chosen = matches[0]
        else:
            try:
                pick = input("\n  Pick a number (or blank to cancel): ").strip()
                if not pick:
                    return
                chosen = matches[int(pick) - 1]
            except (ValueError, IndexError):
                sys.exit("  Invalid choice.")
        cfg.setdefault("item_tags", {})[chosen["instance_id"]] = args.tag
        save_cfg(cfg)
        print(f"  ✓ Tagged {chosen['name']!r} as {args.tag}")


def cmd_id(args, cfg):
    if not args.id:
        sys.exit("--id required")
    tags = cfg.setdefault("item_tags", {})
    if args.untag:
        if args.id in tags:
            del tags[args.id]
            save_cfg(cfg)
            print(f"  ✓ Untagged {args.id}")
        else:
            print(f"  No tag for {args.id}")
    else:
        if args.tag not in VALID_TAGS:
            sys.exit(f"  Invalid tag. Use one of: {', '.join(VALID_TAGS)}")
        tags[args.id] = args.tag
        save_cfg(cfg)
        print(f"  ✓ Tagged {args.id} as {args.tag}")


def cmd_list(cfg):
    tags = cfg.get("item_tags", {})
    if not tags:
        print("  No tagged items.")
        return
    print(f"  {len(tags)} tagged items:")
    # Group by tag
    by_tag = {}
    for iid, tag in tags.items():
        by_tag.setdefault(tag, []).append(iid)
    for tag in VALID_TAGS:
        ids = by_tag.get(tag, [])
        if not ids:
            continue
        print(f"\n    {tag} ({len(ids)}):")
        for iid in ids[:10]:
            print(f"      {iid}")
        if len(ids) > 10:
            print(f"      ... +{len(ids) - 10} more")


def cmd_clear_all(cfg):
    n = len(cfg.get("item_tags", {}))
    if n == 0:
        print("  No tags to clear.")
        return
    ans = input(f"  Clear all {n} tags? [y/N] ").strip().lower()
    if ans != "y":
        return
    cfg["item_tags"] = {}
    save_cfg(cfg)
    print(f"  ✓ Cleared {n} tags.")


def main():
    ap = argparse.ArgumentParser(description="Tag items in your Destiny inventory")
    ap.add_argument("query", nargs="?", help="Search by item name (partial match)")
    ap.add_argument("--id", help="Item instance ID")
    ap.add_argument("--tag", choices=VALID_TAGS,
                    help=f"Tag to apply: {' | '.join(VALID_TAGS)}")
    ap.add_argument("--untag", action="store_true",
                    help="Remove the tag from --id")
    ap.add_argument("--list", action="store_true",
                    help="List all tagged items")
    ap.add_argument("--clear-all", action="store_true",
                    help="Clear ALL tags (asks for confirmation)")
    args = ap.parse_args()

    cfg = load_cfg()

    if args.list:
        cmd_list(cfg)
    elif args.clear_all:
        cmd_clear_all(cfg)
    elif args.id:
        cmd_id(args, cfg)
    elif args.query:
        cmd_search(args, cfg)
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
