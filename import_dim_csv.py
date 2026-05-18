"""
import_dim_csv.py
==================
Import a DIM inventory CSV export into the WISHLIST and INVENTORY sheets.
Lets users who don't want to OAuth get inventory cross-reference for free.

Usage:
    # Import a DIM Weapons CSV (from DIM → Settings → Spreadsheets → Weapons)
    python3 import_dim_csv.py path/to/destinyWeapons.csv

    # Both weapons + armor in one go
    python3 import_dim_csv.py destinyWeapons.csv destinyArmor.csv

DIM CSV format (current as of 2026):
  - First row is the header
  - Common columns: Name, Hash, Id, Tag, Tier, Type, Source, Equippable, Power,
    Element, Masterwork Type, Masterwork Tier, Notes, Perks 0-N, ...
  - For armor: Mobility/Resilience/Recovery/Discipline/Intellect/Strength
    OR (post-EoF) Health/Melee/Grenade/Super/Class/Weapons

The Tag column maps directly to our tag system (favorite/keep/infuse/junk/archive).
"""

import argparse
import csv
import sys
from pathlib import Path

CONFIG_PATH = Path("user_config.json")
import json
import time

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def setc(ws, ref, val, *, bold=False, size=11, color="000000",
         fill=None, italic=False, align="left", wrap=True):
    c = ws[ref]
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER


def load_cfg():
    if not CONFIG_PATH.exists():
        sys.exit("ERROR: user_config.json not found.")
    return json.loads(CONFIG_PATH.read_text())


def parse_csv(path):
    """Read a DIM CSV → list of dict rows."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def normalize_row(row, source_file):
    """Map a DIM CSV row → our unified item schema."""
    # DIM exports vary by content type. Use .get() liberally.
    return {
        "instance_id": row.get("Id") or "",
        "name": row.get("Name") or "",
        "tier": row.get("Tier") or "",
        "type": row.get("Type") or "",
        "element": row.get("Element") or "",
        "slot": row.get("Equippable") or row.get("Equipable") or "",
        "power": row.get("Power") or "",
        "tag": (row.get("Tag") or "").lower(),
        "source_file": Path(source_file).name,
        "perks": " · ".join(
            row.get(f"Perks {i}", "") for i in range(20)
            if row.get(f"Perks {i}")
        ),
        "notes": row.get("Notes") or "",
        "owner": row.get("Owner") or "",
    }


def write_inventory_sheet(workbook_path, items, source_label):
    wb = load_workbook(workbook_path)
    if "INVENTORY" in wb.sheetnames:
        del wb["INVENTORY"]
    ws = wb.create_sheet("INVENTORY", 4)
    widths = [4, 28, 12, 12, 22, 16, 8, 10, 32, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:J1")
    setc(ws, "A1",
         f"  INVENTORY — {len(items):,} items  "
         f"(imported from {source_label} on {time.strftime('%Y-%m-%d %H:%M')})",
         bold=True, size=14, color="FFFFFF", fill="1F2937")
    ws.row_dimensions[1].height = 26

    headers = ["#", "Name", "Tier", "Element", "Type", "Slot", "Power",
               "Tag", "Perks", "Owner"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}3", h, bold=True, fill="F3F4F6",
             align="center" if h in ("#", "Power", "Tag") else "left")

    tag_color = {
        "favorite": "FCD34D",
        "keep":     "A7F3D0",
        "infuse":   "FDE68A",
        "junk":     "FCA5A5",
        "archive":  "E5E7EB",
    }
    row = 4
    for i, it in enumerate(items, 1):
        setc(ws, f"A{row}", i, align="center", size=9, color="6B7280")
        setc(ws, f"B{row}", it["name"], bold=(it["tier"] == "Exotic"))
        setc(ws, f"C{row}", it["tier"], size=10,
             color="F59E0B" if it["tier"] == "Exotic" else "000000")
        setc(ws, f"D{row}", it["element"], size=10)
        setc(ws, f"E{row}", it["type"], size=10)
        setc(ws, f"F{row}", it["slot"], size=10)
        setc(ws, f"G{row}", it["power"], align="center", size=10)
        setc(ws, f"H{row}", it["tag"], align="center", size=10,
             bold=bool(it["tag"]), fill=tag_color.get(it["tag"]))
        setc(ws, f"I{row}", it["perks"], size=9, color="4B5563")
        setc(ws, f"J{row}", it["owner"], size=9, color="6B7280")
        row += 1

    ws.freeze_panes = "A4"
    wb.save(workbook_path)
    print(f"  ✓ Wrote INVENTORY sheet ({len(items):,} rows)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("csvs", nargs="+", help="One or more DIM CSV exports")
    args = ap.parse_args()

    cfg = load_cfg()
    workbook_path = cfg.get("workbook_path", "./my_loadouts.xlsx")
    if not Path(workbook_path).exists():
        sys.exit(f"ERROR: workbook not found at {workbook_path}")

    all_items = []
    sources = []
    for csv_path in args.csvs:
        if not Path(csv_path).exists():
            print(f"  ERROR: {csv_path} not found, skipping")
            continue
        print(f"  Reading {csv_path}...")
        rows = parse_csv(csv_path)
        for row in rows:
            all_items.append(normalize_row(row, csv_path))
        sources.append(Path(csv_path).name)
        print(f"    + {len(rows)} rows")

    # Also pull tags out and save to user_config so they survive across re-fetches
    tags = cfg.setdefault("item_tags", {})
    tag_count = 0
    for it in all_items:
        if it["tag"] and it["instance_id"]:
            tags[it["instance_id"]] = it["tag"]
            tag_count += 1
    if tag_count:
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2) + "\n")
        print(f"  ✓ Imported {tag_count} tags from CSV into user_config.json")

    write_inventory_sheet(workbook_path, all_items, " + ".join(sources))


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n  Aborted.")
        sys.exit(130)
